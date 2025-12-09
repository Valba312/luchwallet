from pathlib import Path

from fastapi import FastAPI, Depends, HTTPException, Header, UploadFile, File
from fastapi.responses import HTMLResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from datetime import datetime, timedelta
import io
import json
import os
from typing import List, Optional
from pydantic import BaseModel, Field

from sqlalchemy import (
    create_engine,
    Column,
    Integer,
    String,
    Text,
    DateTime,
    Boolean,
    text,
)
from sqlalchemy.orm import sessionmaker, declarative_base, Session

from passlib.context import CryptContext
from openpyxl import Workbook
from fastapi.middleware.cors import CORSMiddleware


# ===============================
#   ПУТИ / НАСТРОЙКИ
# ===============================

BASE_DIR = Path(__file__).resolve().parent
CARD_DIST = BASE_DIR / "card" / "dist"

DB_PATH = BASE_DIR / "luchwallet.db"
DATABASE_URL = f"sqlite:///{DB_PATH.as_posix()}"

print("USING DB:", DB_PATH)

PHOTOS_DIR = BASE_DIR / "photos"
os.makedirs(PHOTOS_DIR, exist_ok=True)


# ===============================
#   НАСТРОЙКА БАЗЫ ДАННЫХ
# ===============================

engine = create_engine(
    DATABASE_URL,
    connect_args={"check_same_thread": False} if DATABASE_URL.startswith("sqlite") else {},
)

SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)
Base = declarative_base()

pwd_context = CryptContext(schemes=["argon2"], deprecated="auto")


def get_password_hash(password: str) -> str:
    return pwd_context.hash(password)


def verify_password(plain_password: str, hashed_password: str) -> bool:
    return pwd_context.verify(plain_password, hashed_password)


def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()


# ===============================
#            МОДЕЛИ БД
# ===============================


class Employee(Base):
    __tablename__ = "employees"

    id = Column(Integer, primary_key=True, index=True)
    login = Column(String(50), unique=True, index=True, nullable=False)
    password_hash = Column(String(255), nullable=False)
    # открытый пароль только для админ-панели по ТЗ
    password_plain = Column(String(255), nullable=True)

    initials = Column(String(8), nullable=False)
    name = Column(String(255), nullable=False)
    position = Column(String(255), nullable=False)

    # Привязка к складу/точке
    warehouse = Column(String(255), nullable=True)          # например: "Челябинск · Склад №1"

    # Роль на смене
    shift_role = Column(String(50), nullable=True)          # "receiver" / "loader"

    # Выходит ли на смену сегодня
    on_shift = Column(Boolean, default=False)

    # Ставка за смену (в рублях)
    shift_rate = Column(Integer, nullable=True)

    # Оклад
    rate = Column(String(50), nullable=True)
    experience = Column(String(50), nullable=True)
    status = Column(String(100), nullable=True)

    # Отображаемые поля
    salary = Column(String(50), nullable=True)   # отображаемый баланс (строка)
    hours = Column(String(50), nullable=True)    # отображаемые часы за месяц
    hours_detail = Column(String(255), nullable=True)

    penalties_json = Column(Text, nullable=True)
    absences_json = Column(Text, nullable=True)

    error_text = Column(String(255), nullable=True)

    photo_url = Column(String(512), nullable=True)

    # === динамический баланс ===
    balance_int = Column(Integer, nullable=True)               # фактический баланс в рублях
    contract_hours_per_month = Column(Integer, nullable=True)  # нормочасы в месяц
    hourly_rate = Column(Integer, nullable=True)               # ₽/час
    schedule_type = Column(String(50), nullable=True)          # office / None / ...
    work_start_hour = Column(Integer, nullable=True)           # 0–23
    work_end_hour = Column(Integer, nullable=True)             # 0–23 (не включая)
    last_balance_update = Column(DateTime, nullable=True)

    is_active = Column(Boolean, default=True)
    created_at = Column(DateTime, default=datetime.utcnow)


class Admin(Base):
    __tablename__ = "admins"

    id = Column(Integer, primary_key=True, index=True)
    login = Column(String(50), unique=True, index=True, nullable=False)
    password_hash = Column(String(255), nullable=False)
    name = Column(String(255), nullable=False)
    created_at = Column(DateTime, default=datetime.utcnow)


class Payment(Base):
    """
    История начислений/удержаний сотрудника.
    amount в рублях:
      >0 — начисление
      <0 — удержание
    """
    __tablename__ = "payments"

    id = Column(Integer, primary_key=True, index=True)
    employee_id = Column(Integer, index=True, nullable=False)
    type = Column(String(50), nullable=False)  # salary / bonus / overtime / night / fine / other
    amount = Column(Integer, nullable=False)
    comment = Column(String(255), nullable=True)
    created_at = Column(DateTime, default=datetime.utcnow, index=True)


class EmployeeMonthStat(Base):
    """
    Помесячная статистика дохода/часов/штрафов.
    Хранится отдельно от основной карточки сотрудника.
    """
    __tablename__ = "employee_month_stats"

    id = Column(Integer, primary_key=True, index=True)
    employee_id = Column(Integer, index=True, nullable=False)
    year = Column(Integer, nullable=False)
    month = Column(Integer, nullable=False)  # 1-12
    month_key = Column(String(8), nullable=False)  # jan, feb, ...

    income = Column(Integer, nullable=False, default=0)  # общий доход за месяц
    salary = Column(Integer, nullable=True)              # начисленная зарплата
    hours = Column(Integer, nullable=True)               # отработанные часы

    penalties_json = Column(Text, nullable=True)         # JSON-список строк
    absences_json = Column(Text, nullable=True)          # JSON-список строк

    created_at = Column(DateTime, default=datetime.utcnow, index=True)


# ===============================
#         Pydantic-схемы
# ===============================


class LoginRequest(BaseModel):
    role: str   # "employee" или "admin"/"manager"
    login: str
    password: str


class LoginResponse(BaseModel):
    role: str
    login: str
    data: dict


class EmployeeBase(BaseModel):
    initials: str
    name: str
    position: str
    rate: Optional[str] = None
    experience: Optional[str] = None
    status: Optional[str] = None
    salary: Optional[str] = None
    hours: Optional[str] = None
    hours_detail: Optional[str] = None
    penalties: List[str] = []
    absences: List[str] = []
    error_text: Optional[str] = ""
    photo_url: Optional[str] = None
    warehouse: Optional[str] = None
    shift_role: Optional[str] = None     # "receiver" / "loader"
    on_shift: Optional[bool] = False
    shift_rate: Optional[int] = None

class EmployeeCreate(EmployeeBase):
    login: str
    password: str


class EmployeeUpdate(BaseModel):
    initials: Optional[str] = None
    name: Optional[str] = None
    position: Optional[str] = None
    rate: Optional[str] = None
    experience: Optional[str] = None
    status: Optional[str] = None
    salary: Optional[str] = None
    hours: Optional[str] = None
    hours_detail: Optional[str] = None
    penalties: Optional[List[str]] = None
    absences: Optional[List[str]] = None
    error_text: Optional[str] = None
    photo_url: Optional[str] = None
    # смена логина/пароля
    login: Optional[str] = None
    password: Optional[str] = None
    warehouse: Optional[str] = None
    shift_role: Optional[str] = None
    on_shift: Optional[bool] = None
    shift_rate: Optional[int] = None


class EmployeeShort(BaseModel):
    id: int
    login: str
    name: str
    position: str
    is_active: bool
    photo_url: Optional[str] = None

    warehouse: Optional[str] = None
    shift_role: Optional[str] = None
    on_shift: Optional[bool] = False

    password: Optional[str] = Field(None, alias="password_plain")

    class Config:
        orm_mode = True
        allow_population_by_field_name = True


class EmployeeDetail(EmployeeBase):
    id: int
    login: str
    is_active: bool
    password: Optional[str] = Field(None, alias="password_plain")

    class Config:
        orm_mode = True
        allow_population_by_field_name = True


class PaymentBaseSchema(BaseModel):
    type: str
    amount: int
    comment: Optional[str] = None


class PaymentCreate(PaymentBaseSchema):
    pass


class PaymentOut(PaymentBaseSchema):
    id: int
    employee_id: int
    created_at: datetime

    class Config:
        orm_mode = True


class EmployeeSelfPaymentsRequest(BaseModel):
    """
    Для /api/employee/payments — запрос с логином и паролем сотрудника.
    (без поля role, как и шлёт фронт)
    """
    login: str
    password: str


# ===============================
#        ВСПОМОГАТЕЛЬНОЕ
# ===============================


def json_dumps_list(values: List[str]) -> str:
    return json.dumps(values, ensure_ascii=False)


def json_loads_list(raw: Optional[str]) -> List[str]:
    if not raw:
        return []
    try:
        return json.loads(raw)
    except Exception:
        return []


def money_to_int(value: Optional[str]) -> int:
    """'92 430 ₽' -> 92430"""
    if not value:
        return 0
    digits = "".join(ch for ch in str(value) if ch.isdigit())
    return int(digits) if digits else 0


def int_to_money(value: int) -> str:
    """92430 -> '92 430 ₽'"""
    s = f"{value:,}".replace(",", " ")
    return f"{s} ₽"


def is_office_work_time(dt: datetime, start_hour: int, end_hour: int) -> bool:
    """
    Простейшая модель: офисники работают Пн–Пт с start_hour до end_hour (без учёта обеда).
    """
    if dt.weekday() >= 5:  # 5,6 = сб, вс
        return False
    h = dt.hour
    if start_hour <= end_hour:
        return start_hour <= h < end_hour
    # если смена «через ночь»
    return h >= start_hour or h < end_hour


def ensure_emp_balance_initialized(emp: Employee):
    """
    Гарантируем, что balance_int и salary синхронизированы.
    """
    if emp.balance_int is None:
        emp.balance_int = money_to_int(emp.salary or "0")
    if emp.salary is None:
        emp.salary = int_to_money(emp.balance_int or 0)


def accrue_balance_for_employee(emp: Employee, now: Optional[datetime] = None):
    """
    Обновляет баланс сотрудника в БД по почасовой ставке.
    Для простоты:
      - автоматом считаем только для schedule_type == 'office'
      - берём каждый полный час между last_balance_update и now,
        если час попадает в рабочее время – начисляем hourly_rate.
    """
    if emp.schedule_type != "office":
        return
    if not emp.hourly_rate:
        return

    if now is None:
        now = datetime.utcnow()

    # инициализация
    if not emp.last_balance_update:
        emp.last_balance_update = now
        ensure_emp_balance_initialized(emp)
        return

    cursor = emp.last_balance_update.replace(minute=0, second=0, microsecond=0)
    if cursor >= now:
        return

    hours_to_pay = 0
    while cursor + timedelta(hours=1) <= now:
        if is_office_work_time(cursor, emp.work_start_hour or 8, emp.work_end_hour or 19):
            hours_to_pay += 1
        cursor += timedelta(hours=1)

    if hours_to_pay > 0:
        ensure_emp_balance_initialized(emp)
        emp.balance_int += hours_to_pay * emp.hourly_rate
        emp.salary = int_to_money(emp.balance_int)

    emp.last_balance_update = cursor


MONTH_META = {
    1: {"key": "jan", "short": "Янв", "full": "Январь"},
    2: {"key": "feb", "short": "Фев", "full": "Февраль"},
    3: {"key": "mar", "short": "Мар", "full": "Март"},
    4: {"key": "apr", "short": "Апр", "full": "Апрель"},
    5: {"key": "may", "short": "Май", "full": "Май"},
    6: {"key": "jun", "short": "Июн", "full": "Июнь"},
    7: {"key": "jul", "short": "Июл", "full": "Июль"},
    8: {"key": "aug", "short": "Авг", "full": "Август"},
    9: {"key": "sep", "short": "Сен", "full": "Сентябрь"},
    10: {"key": "oct", "short": "Окт", "full": "Октябрь"},
    11: {"key": "nov", "short": "Ноя", "full": "Ноябрь"},
    12: {"key": "dec", "short": "Дек", "full": "Декабрь"},
}


def build_months_for_employee(db: Session, emp_id: int) -> List[dict]:
    """Отдаём месяцы в удобном для фронта формате."""
    stats = (
        db.query(EmployeeMonthStat)
        .filter(EmployeeMonthStat.employee_id == emp_id)
        .order_by(EmployeeMonthStat.year.asc(), EmployeeMonthStat.month.asc())
        .all()
    )
    result: List[dict] = []
    for s in stats:
        meta = MONTH_META.get(s.month, None)
        if meta:
            key = s.month_key or meta["key"]
            short = meta["short"]
            full = meta["full"]
        else:
            key = s.month_key or str(s.month)
            short = key
            full = key
        result.append(
            {
                "key": key,
                "short": short,
                "fullName": full,
                "year": s.year,
                "month": s.month,
                "income": s.income,
                "salary": s.salary,
                "hours": s.hours,
                "penalties": json_loads_list(s.penalties_json),
                "absences": json_loads_list(s.absences_json),
            }
        )
    return result


def seed_month_stats_for_employee(
    db: Session,
    emp: Employee,
    months_data: List[dict],
):
    """Создаём помесячную статистику, если её ещё нет."""
    if not emp:
        return
    exists = (
        db.query(EmployeeMonthStat)
        .filter(EmployeeMonthStat.employee_id == emp.id)
        .count()
    )
    if exists:
        return

    for m in months_data:
        stat = EmployeeMonthStat(
            employee_id=emp.id,
            year=m["year"],
            month=m["month"],
            month_key=m["key"],
            income=m["income"],
            salary=m.get("salary"),
            hours=m.get("hours"),
            penalties_json=json_dumps_list(m.get("penalties", [])),
            absences_json=json_dumps_list(m.get("absences", [])),
        )
        db.add(stat)


def update_month_stat_on_payment(
    db: Session,
    emp_id: int,
    amount_diff: int,
    created_at: datetime,
    payment_type: str,
    comment: Optional[str] = None,
    reverse: bool = False,
):
    """
    Обновляет EmployeeMonthStat при добавлении/удалении платежа.
    amount_diff — сумма платежа (для удаления можно использовать ту же сумму, но reverse=True).
    reverse=True — "откатить" операцию (income -= amount_diff вместо +=).
    """
    year = created_at.year
    month = created_at.month

    stat = (
        db.query(EmployeeMonthStat)
        .filter(
            EmployeeMonthStat.employee_id == emp_id,
            EmployeeMonthStat.year == year,
            EmployeeMonthStat.month == month,
        )
        .first()
    )

    if not stat:
        # если удаляем операцию, а стата нет — ничего не делаем
        if reverse:
            return
        meta = MONTH_META.get(month, {"key": str(month)})
        stat = EmployeeMonthStat(
            employee_id=emp_id,
            year=year,
            month=month,
            month_key=meta["key"],
            income=0,
            salary=0,
            hours=None,
            penalties_json=json_dumps_list([]),
            absences_json=json_dumps_list([]),
        )
        db.add(stat)
        db.flush()

    sign = -1 if reverse else 1
    delta = sign * amount_diff

    stat.income = (stat.income or 0) + delta

    # немного логики для salary: учитываем только тип "salary" как "зарплату"
    if payment_type == "salary":
        stat.salary = (stat.salary or 0) + delta


# ===============================
#        ИНИЦИАЛИЗАЦИЯ БД
# ===============================

def init_db():
    """Создаём таблицы и демо-данные."""
    Base.metadata.create_all(bind=engine)
    db = SessionLocal()
    try:
        # добавляем колонку password_plain, если её нет
        try:
            db.execute(text("ALTER TABLE employees ADD COLUMN password_plain VARCHAR(255);"))
            db.commit()
        except Exception:
            db.rollback()

        # --- добавляем новые колонки, если их ещё нет ---
        for ddl in [
            "ALTER TABLE employees ADD COLUMN balance_int INTEGER;",
            "ALTER TABLE employees ADD COLUMN contract_hours_per_month INTEGER;",
            "ALTER TABLE employees ADD COLUMN hourly_rate INTEGER;",
            "ALTER TABLE employees ADD COLUMN schedule_type VARCHAR(50);",
            "ALTER TABLE employees ADD COLUMN work_start_hour INTEGER;",
            "ALTER TABLE employees ADD COLUMN work_end_hour INTEGER;",
            "ALTER TABLE employees ADD COLUMN last_balance_update DATETIME;",
            "ALTER TABLE employees ADD COLUMN warehouse VARCHAR(255);",
            "ALTER TABLE employees ADD COLUMN shift_role VARCHAR(50);",
            "ALTER TABLE employees ADD COLUMN on_shift BOOLEAN DEFAULT 0;",
            "ALTER TABLE employees ADD COLUMN shift_rate INTEGER;",
        ]:
            try:
                db.execute(text(ddl))
                db.commit()
            except Exception:
                db.rollback()

        # демо-сотрудник ivan
        if not db.query(Employee).filter_by(login="ivan").first():
            emp = Employee(
                login="ivan",
                password_hash=get_password_hash("1234"),
                password_plain="1234",
                initials="ИИ",
                name="Иванов Иван Иванович",
                position="Водитель грузового автомобиля · Колонна № 3",
                rate="1 850 ₽/смена",
                experience="4 года 7 мес.",
                status="Активен · Основное место",
                salary="92 430 ₽",
                hours="152 ч",
                hours_detail="Переработка: 18 ч · Ночные: 12 ч.",
                penalties_json=json_dumps_list(
                    [
                        "Штрафов: 1 — превышение времени стоянки",
                        "Прогулы: нет",
                        "Замечания: отсутствуют",
                    ]
                ),
                absences_json=json_dumps_list(
                    [
                        "Больничные: 3 дня (ОРВИ)",
                        "Отпуск: 14/28 дней",
                        "Отсутствия: 1 день за свой счёт",
                    ]
                ),
                error_text="",
                photo_url=None,
                balance_int=92430,
                contract_hours_per_month=152,
                hourly_rate=608,
                schedule_type=None,
                work_start_hour=None,
                work_end_hour=None,
                last_balance_update=datetime.utcnow(),
                warehouse="Челябинск · Склад №1",
                shift_role="loader",      # кладовщик/грузит
                on_shift=True,
                shift_rate=1850,
            )
            db.add(emp)

        # демо-сотрудник anna
        if not db.query(Employee).filter_by(login="anna").first():
            emp = Employee(
                login="anna",
                password_hash=get_password_hash("qwerty"),
                password_plain="qwerty",
                initials="АП",
                name="Антонова Анна Петровна",
                position="Диспетчер смен · Офис № 2",
                rate="2 050 ₽/смена",
                experience="2 года 3 мес.",
                status="Активна · Совместительство",
                salary="74 300 ₽",
                hours="128 ч",
                hours_detail="Переработка: 6 ч · Ночные: 4 ч.",
                penalties_json=json_dumps_list(
                    [
                        "Штрафов: нет",
                        "Прогулы: нет",
                        "Замечания: 1 — опоздание на планёрку",
                    ]
                ),
                absences_json=json_dumps_list(
                    [
                        "Больничные: не было",
                        "Отпуск: 7/28 дней",
                        "Отсутствия: нет",
                    ]
                ),
                error_text="",
                photo_url=None,
                balance_int=74300,
                contract_hours_per_month=128,
                hourly_rate=580,
                schedule_type="office",
                work_start_hour=8,
                work_end_hour=19,
                last_balance_update=datetime.utcnow(),
                warehouse="Челябинск · Склад №1",
                shift_role="receiver",    # приёмщик
                on_shift=False,
                shift_rate=2050,
            )
            db.add(emp)

        # демо-админ admin / admin123
        if not db.query(Admin).filter_by(login="admin").first():
            admin = Admin(
                login="admin",
                password_hash=get_password_hash("admin123"),
                name="Администратор системы",
            )
            db.add(admin)

        # демо-менеджер для карточки сотрудника: manager / 123456
        manager_admin = db.query(Admin).filter_by(login="manager").first()
        if manager_admin:
            manager_admin.password_hash = get_password_hash("123456")
            if not manager_admin.name:
                manager_admin.name = "Менеджер склада (демо)"
        else:
            manager_admin = Admin(
                login="manager",
                password_hash=get_password_hash("123456"),
                name="Менеджер склада (демо)",
            )
            db.add(manager_admin)

        db.flush()

        # помесячные данные по ivan
        ivan = db.query(Employee).filter_by(login="ivan").first()
        if ivan:
            seed_month_stats_for_employee(
                db,
                ivan,
                [
                    {
                        "year": 2024,
                        "month": 6,
                        "key": "jun",
                        "income": 87762,
                        "salary": 87762,
                        "hours": 140,
                        "penalties": ["Штрафов: нет", "Прогулы: нет"],
                        "absences": ["Больничные: 1 день", "Отпуск: 5/28 дней"],
                    },
                    {
                        "year": 2024,
                        "month": 7,
                        "key": "jul",
                        "income": 90789,
                        "salary": 90789,
                        "hours": 144,
                        "penalties": ["Штрафов: 1 — превышение времени стоянки"],
                        "absences": ["Больничные: нет", "Отпуск: 9/28 дней"],
                    },
                    {
                        "year": 2024,
                        "month": 8,
                        "key": "aug",
                        "income": 89562,
                        "salary": 89562,
                        "hours": 146,
                        "penalties": ["Штрафов: нет"],
                        "absences": ["Больничные: нет", "Отпуск: 14/28 дней"],
                    },
                    {
                        "year": 2024,
                        "month": 9,
                        "key": "sep",
                        "income": 90349,
                        "salary": 90349,
                        "hours": 148,
                        "penalties": ["Штрафов: нет"],
                        "absences": ["Отсутствия: 1 день за свой счёт"],
                    },
                    {
                        "year": 2024,
                        "month": 10,
                        "key": "oct",
                        "income": 89967,
                        "salary": 92430,
                        "hours": 152,
                        "penalties": ["Штрафов: 1 — превышение времени стоянки"],
                        "absences": ["Больничные: 3 дня (ОРВИ)", "Отпуск: 14/28 дней"],
                    },
                    {
                        "year": 2024,
                        "month": 11,
                        "key": "nov",
                        "income": 96836,
                        "salary": 102430,
                        "hours": 156,
                        "penalties": ["Штрафов: нет"],
                        "absences": ["Больничные: нет", "Отпуск: 18/28 дней"],
                    },
                ],
            )

        # помесячные данные по anna
        anna = db.query(Employee).filter_by(login="anna").first()
        if anna:
            seed_month_stats_for_employee(
                db,
                anna,
                [
                    {
                        "year": 2024,
                        "month": 6,
                        "key": "jun",
                        "income": 60310,
                        "salary": 60310,
                        "hours": 116,
                        "penalties": [],
                        "absences": ["Больничные: нет"],
                    },
                    {
                        "year": 2024,
                        "month": 7,
                        "key": "jul",
                        "income": 62140,
                        "salary": 62140,
                        "hours": 120,
                        "penalties": [],
                        "absences": ["Отпуск: 3/28 дней"],
                    },
                    {
                        "year": 2024,
                        "month": 8,
                        "key": "aug",
                        "income": 64300,
                        "salary": 64300,
                        "hours": 124,
                        "penalties": [],
                        "absences": ["Отпуск: 7/28 дней"],
                    },
                    {
                        "year": 2024,
                        "month": 9,
                        "key": "sep",
                        "income": 70100,
                        "salary": 70100,
                        "hours": 128,
                        "penalties": ["Замечания: 1 — опоздание на планёрку"],
                        "absences": [],
                    },
                    {
                        "year": 2024,
                        "month": 10,
                        "key": "oct",
                        "income": 71500,
                        "salary": 71500,
                        "hours": 130,
                        "penalties": [],
                        "absences": [],
                    },
                    {
                        "year": 2024,
                        "month": 11,
                        "key": "nov",
                        "income": 76800,
                        "salary": 76800,
                        "hours": 132,
                        "penalties": [],
                        "absences": ["Отпуск: 10/28 дней"],
                    },
                ],
            )

        db.commit()
    finally:
        db.close()


# ===============================
#     ПРОВЕРКА ПРАВ АДМИНА
# ===============================

def require_admin(
    db: Session = Depends(get_db),
    admin_login: str = Header(..., alias="X-Admin-Login"),
    admin_password: str = Header(..., alias="X-Admin-Password"),
) -> Admin:
    login_value = admin_login.lower()

    # Спец-случай для менеджера карточки:
    # любые заголовки X-Admin-Login: manager / любой пароль считаем валидными
    if login_value == "manager":
        adm = db.query(Admin).filter(Admin.login == "manager").first()
        if not adm:
            adm = Admin(
                login="manager",
                password_hash=get_password_hash("123456"),
                name="Менеджер склада (демо)",
            )
            db.add(adm)
            db.commit()
            db.refresh(adm)
        return adm

    # Обычные админы — по старой схеме
    adm = db.query(Admin).filter(Admin.login == login_value).first()
    if not adm or not verify_password(admin_password, adm.password_hash):
        raise HTTPException(status_code=401, detail="Админ не авторизован")
    return adm


# ===============================
#           FASTAPI APP
# ===============================

app = FastAPI(title="LuchWallet API", version="2.3.0")

# фотки сотрудников
app.mount("/static", StaticFiles(directory=str(PHOTOS_DIR)), name="static")

# статика Vite-карточки (js/css) — /assets/...
if (CARD_DIST / "assets").exists():
    app.mount(
        "/assets",
        StaticFiles(directory=str(CARD_DIST / "assets")),
        name="card_assets",
    )

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.on_event("startup")
def on_startup():
    init_db()


@app.get("/api/health")
def health():
    return {"status": "ok", "app": "LuchWallet API"}


# ---------- ЛОГИН (сотрудник / админ / менеджер) ----------

@app.post("/api/login", response_model=LoginResponse)
def login_endpoint(payload: LoginRequest, db: Session = Depends(get_db)):
    role = (payload.role or "").lower()
    login_value = payload.login.strip().lower()
    password = payload.password

    # ===== ОТДЕЛЬНЫЙ СЛУЧАЙ ДЛЯ КАРТОЧКИ СОТРУДНИКА (manager) =====
    if login_value == "manager":
        # гарантируем, что в БД есть такой админ
        adm = db.query(Admin).filter(Admin.login == "manager").first()
        if not adm:
            adm = Admin(
                login="manager",
                password_hash=get_password_hash("123456"),
                name="Менеджер склада (демо)",
            )
            db.add(adm)
            db.commit()
            db.refresh(adm)

        # Для демо НЕ проверяем пароль — всегда успешный вход
        return LoginResponse(
            role="manager",
            login=login_value,
            data={"name": adm.name},
        )

    # ===== ЛОГИН СОТРУДНИКА (кошелёк) =====
    if role == "employee":
        emp = db.query(Employee).filter(Employee.login == login_value).first()
        if not emp or not verify_password(password, emp.password_hash):
            raise HTTPException(status_code=401, detail="Неверный логин или пароль")

        accrue_balance_for_employee(emp)
        db.commit()
        db.refresh(emp)

        months = build_months_for_employee(db, emp.id)

        data = {
            "id": emp.id,
            "initials": emp.initials,
            "name": emp.name,
            "position": emp.position,
            "rate": emp.rate,
            "experience": emp.experience,
            "status": emp.status,
            "salary": emp.salary,
            "hours": emp.hours,
            "hoursDetail": emp.hours_detail,
            "penalties": json_loads_list(emp.penalties_json),
            "absences": json_loads_list(emp.absences_json),
            "errorText": emp.error_text or "",
            "photo_url": emp.photo_url,
            "months": months,
        }

        return LoginResponse(role="employee", login=login_value, data=data)

    # ===== ЛОГИН ОБЫЧНОГО АДМИНА =====
    adm = db.query(Admin).filter(Admin.login == login_value).first()
    if not adm or not verify_password(password, adm.password_hash):
        raise HTTPException(status_code=401, detail="Неверный логин или пароль")

    return LoginResponse(
        role="admin",
        login=login_value,
        data={"name": adm.name},
    )


# ---------- СПИСОК СОТРУДНИКОВ ДЛЯ АДМИНА ----------

@app.get("/api/employees", response_model=List[EmployeeShort])
def list_employees(admin: Admin = Depends(require_admin), db: Session = Depends(get_db)):
    employees = (
        db.query(Employee)
        .filter(Employee.is_active == True)
        .order_by(Employee.id.asc())
        .all()
    )
    return employees


@app.get("/api/employees/{employee_id}", response_model=EmployeeDetail)
def get_employee(
    employee_id: int,
    admin: Admin = Depends(require_admin),
    db: Session = Depends(get_db),
):
    emp = db.query(Employee).filter(Employee.id == employee_id).first()
    if not emp:
        raise HTTPException(status_code=404, detail="Сотрудник не найден")
    penalties = json_loads_list(emp.penalties_json)
    absences = json_loads_list(emp.absences_json)
    return EmployeeDetail(
        id=emp.id,
        login=emp.login,
        initials=emp.initials,
        name=emp.name,
        position=emp.position,
        rate=emp.rate,
        experience=emp.experience,
        status=emp.status,
        salary=emp.salary,
        hours=emp.hours,
        hours_detail=emp.hours_detail,
        penalties=penalties,
        absences=absences,
        error_text=emp.error_text,
        photo_url=emp.photo_url,
        is_active=emp.is_active,
        password=emp.password_plain,
    )


@app.post("/api/employees", response_model=EmployeeDetail)
def create_employee(
    payload: EmployeeCreate,
    admin: Admin = Depends(require_admin),
    db: Session = Depends(get_db),
):
    if db.query(Employee).filter(Employee.login == payload.login.lower()).first():
        raise HTTPException(status_code=400, detail="Логин уже занят")

    emp = Employee(
        login=payload.login.lower(),
        password_hash=get_password_hash(payload.password),
        password_plain=payload.password,
        initials=payload.initials,
        name=payload.name,
        position=payload.position,
        rate=payload.rate,
        experience=payload.experience,
        status=payload.status,
        salary=payload.salary,
        hours=payload.hours,
        hours_detail=payload.hours_detail,
        penalties_json=json_dumps_list(payload.penalties or []),
        absences_json=json_dumps_list(payload.absences or []),
        error_text=payload.error_text,
        photo_url=payload.photo_url,

        warehouse=payload.warehouse,
        shift_role=payload.shift_role,
        on_shift=payload.on_shift,
        shift_rate=payload.shift_rate,

        is_active=True,
    )


    # инициализируем динамический баланс и нормочасы
    emp.balance_int = money_to_int(payload.salary or "0")
    try:
        hours_int = int("".join(ch for ch in (payload.hours or "") if ch.isdigit()))
    except ValueError:
        hours_int = None
    emp.contract_hours_per_month = hours_int
    if hours_int and hours_int > 0:
        emp.hourly_rate = emp.balance_int // hours_int
    else:
        emp.hourly_rate = None

    db.add(emp)
    db.commit()
    db.refresh(emp)

    return get_employee(emp.id, admin=admin, db=db)


@app.put("/api/employees/{employee_id}", response_model=EmployeeDetail)
def update_employee(
    employee_id: int,
    payload: EmployeeUpdate,
    admin: Admin = Depends(require_admin),
    db: Session = Depends(get_db),
):
    emp = db.query(Employee).filter(Employee.id == employee_id).first()
    if not emp:
        raise HTTPException(status_code=404, detail="Сотрудник не найден")

    if payload.login:
        new_login = payload.login.lower()
        if new_login != emp.login and db.query(Employee).filter(Employee.login == new_login).first():
            raise HTTPException(status_code=400, detail="Логин уже занят")
        emp.login = new_login

    if payload.password:
        emp.password_hash = get_password_hash(payload.password)
        emp.password_plain = payload.password

    for field in [
        "initials",
        "name",
        "position",
        "rate",
        "experience",
        "status",
        "salary",
        "hours",
        "hours_detail",
        "error_text",
        "photo_url",
        "warehouse",
        "shift_role",
        "on_shift",
        "shift_rate",
    ]:
        val = getattr(payload, field)
        if val is not None:
            setattr(emp, field, val)

    if payload.penalties is not None:
        emp.penalties_json = json_dumps_list(payload.penalties)
    if payload.absences is not None:
        emp.absences_json = json_dumps_list(payload.absences)

    # пересчёт баланс_int и нормочасов, если пришли salary/hours
    if payload.salary is not None:
        emp.balance_int = money_to_int(emp.salary)
    if payload.hours is not None:
        try:
            hours_int = int("".join(ch for ch in (emp.hours or "") if ch.isdigit()))
        except ValueError:
            hours_int = None
        emp.contract_hours_per_month = hours_int
        if hours_int and hours_int > 0 and emp.balance_int is not None:
            emp.hourly_rate = emp.balance_int // hours_int

    db.commit()
    db.refresh(emp)

    return get_employee(emp.id, admin=admin, db=db)


@app.delete("/api/employees/{employee_id}")
def delete_employee(
    employee_id: int,
    admin: Admin = Depends(require_admin),
    db: Session = Depends(get_db),
):
    emp = db.query(Employee).filter(Employee.id == employee_id).first()
    if not emp:
        raise HTTPException(status_code=404, detail="Сотрудник не найден")
    # мягкое удаление
    emp.is_active = False
    db.commit()
    return {"status": "ok", "id": employee_id}


# ---------- ФОТО СОТРУДНИКА ----------

@app.post("/api/employees/{employee_id}/photo")
def upload_employee_photo(
    employee_id: int,
    file: UploadFile = File(...),
    admin: Admin = Depends(require_admin),
    db: Session = Depends(get_db),
):
    emp = db.query(Employee).filter(Employee.id == employee_id).first()
    if not emp:
        raise HTTPException(status_code=404, detail="Сотрудник не найден")

    ext = ""
    if "." in file.filename:
        ext = "." + file.filename.rsplit(".", 1)[-1].lower()
    filename = f"emp_{employee_id}{ext}"
    filepath = PHOTOS_DIR / filename

    with filepath.open("wb") as f:
        f.write(file.file.read())

    emp.photo_url = f"/static/{filename}"
    db.commit()
    db.refresh(emp)

    return {"photo_url": emp.photo_url}


# ---------- ЭКСПОРТ КАРТОЧКИ СОТРУДНИКА В EXCEL ----------

@app.get("/api/employees/{employee_id}/export")
def export_employee_excel(
    employee_id: int,
    admin: Admin = Depends(require_admin),
    db: Session = Depends(get_db),
):
    emp = db.query(Employee).filter(Employee.id == employee_id).first()
    if not emp:
        raise HTTPException(status_code=404, detail="Сотрудник не найден")

    wb = Workbook()
    ws = wb.active
    ws.title = "Карточка сотрудника"

    penalties = json_loads_list(emp.penalties_json)
    absences = json_loads_list(emp.absences_json)

    data_rows = [
        ("ФИО", emp.name),
        ("Логин", emp.login),
        ("Должность", emp.position),
        ("Оклад", emp.rate),
        ("Стаж", emp.experience),
        ("Статус", emp.status),
        ("Баланс", emp.salary),
        ("Отработанное время", emp.hours),
        ("Детализация времени", emp.hours_detail),
        ("Штрафы и дисциплина", "\n".join(penalties)),
        ("Больничные и отсутствия", "\n".join(absences)),
        ("Примечание / ошибка", emp.error_text),
    ]

    row_idx = 1
    for label, value in data_rows:
        ws.cell(row=row_idx, column=1, value=label)
        ws.cell(row=row_idx, column=2, value=value)
        row_idx += 1

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)

    filename = f"employee_{employee_id}_card.xlsx"

    return StreamingResponse(
        bio,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"'
        },
    )


# ---------- ПЛАТЕЖИ / НАЧИСЛЕНИЯ ДЛЯ АДМИНА ----------

@app.get("/api/employees/{employee_id}/payments", response_model=List[PaymentOut])
def list_payments_for_employee(
    employee_id: int,
    admin: Admin = Depends(require_admin),
    db: Session = Depends(get_db),
):
    emp = db.query(Employee).filter(Employee.id == employee_id).first()
    if not emp:
        raise HTTPException(status_code=404, detail="Сотрудник не найден")

    payments = (
        db.query(Payment)
        .filter(Payment.employee_id == employee_id)
        .order_by(Payment.created_at.desc(), Payment.id.desc())
        .all()
    )
    return payments


@app.post("/api/employees/{employee_id}/payments", response_model=PaymentOut)
def create_payment_for_employee(
    employee_id: int,
    payload: PaymentCreate,
    admin: Admin = Depends(require_admin),
    db: Session = Depends(get_db),
):
    emp = db.query(Employee).filter(Employee.id == employee_id).first()
    if not emp:
        raise HTTPException(status_code=404, detail="Сотрудник не найден")

    ensure_emp_balance_initialized(emp)
    emp.balance_int += payload.amount
    emp.salary = int_to_money(emp.balance_int)

    payment = Payment(
        employee_id=employee_id,
        type=payload.type,
        amount=payload.amount,
        comment=payload.comment,
    )
    db.add(payment)
    db.flush()

    update_month_stat_on_payment(
        db=db,
        emp_id=employee_id,
        amount_diff=payload.amount,
        created_at=payment.created_at,
        payment_type=payload.type,
        comment=payload.comment,
        reverse=False,
    )

    db.commit()
    db.refresh(payment)

    return payment


@app.delete("/api/employees/{employee_id}/payments/{payment_id}")
def delete_payment_for_employee(
    employee_id: int,
    payment_id: int,
    admin: Admin = Depends(require_admin),
    db: Session = Depends(get_db),
):
    payment = db.query(Payment).filter(Payment.id == payment_id).first()
    if not payment:
        raise HTTPException(status_code=404, detail="Платёж не найден")

    emp = db.query(Employee).filter(Employee.id == payment.employee_id).first()
    if emp:
        ensure_emp_balance_initialized(emp)
        emp.balance_int -= payment.amount
        emp.salary = int_to_money(emp.balance_int)

        update_month_stat_on_payment(
            db=db,
            emp_id=emp.id,
            amount_diff=payment.amount,
            created_at=payment.created_at,
            payment_type=payment.type,
            comment=payment.comment,
            reverse=True,
        )

    db.delete(payment)
    db.commit()
    return {"status": "deleted", "id": payment_id}


# ---------- СОТРУДНИК: СВОЯ ИСТОРИЯ ОПЕРАЦИЙ (баланс) ----------

@app.post("/api/employee/payments", response_model=List[PaymentOut])
def list_payments_for_employee_self(
    payload: EmployeeSelfPaymentsRequest,
    db: Session = Depends(get_db),
):
    """
    Эндпоинт для фронта при клике на баланс (модалка истории).
    На вход: login + password сотрудника.
    """
    login_value = payload.login.strip().lower()
    emp = db.query(Employee).filter(Employee.login == login_value).first()
    if not emp or not verify_password(payload.password, emp.password_hash):
        raise HTTPException(status_code=401, detail="Неверный логин или пароль")

    accrue_balance_for_employee(emp)
    db.commit()

    payments = (
        db.query(Payment)
        .filter(Payment.employee_id == emp.id)
        .order_by(Payment.created_at.desc(), Payment.id.desc())
        .all()
    )
    return payments


# ---------- ФРОНТ: КАРТОЧКА СОТРУДНИКА (React/Vite) ----------

@app.get("/card", response_class=HTMLResponse)
def card_page():
    """
    Карточка сотрудника по адресу: http://127.0.0.1:8000/card
    """
    index_file = CARD_DIST / "index.html"
    if not index_file.exists():
        raise HTTPException(
            status_code=500,
            detail="Карточка сотрудника не собрана. В папке card выполни `npm run build`.",
        )
    return HTMLResponse(index_file.read_text(encoding="utf-8"))
app.mount(
    "/",  # корень сайта
    StaticFiles(directory="wallet", html=True),
    name="wallet",
)
